#xls
#openpyxl --> xlsx
from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    #打開文件
    file = load_workbook(filename=path)
    #所有表格的名稱
    #print(file.get_sheet_names())
    sheets = file.sheetnames
    print(len(sheets))
    for sheetName in sheets:
        sheet = file[sheetName]
        sheetInfo = []
        for lineNum in range(1, sheet.max_row + 1):
            linList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                linList.append(value)
            sheetInfo.append(linList)
        #將一張表的數據存到字典
        dic[sheetName] = sheetInfo
    return dic
#不能處理Xls文件
path = r"E:\code\python\1.xlsx"
dic = readXlsxFile(path)
#取所有表
print(dic)
#取一張表
#print(dic["第一張表"])
print(len(dic))