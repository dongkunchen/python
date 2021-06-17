#xls
#openpyxl --> xlsx
from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打開文件
    file = load_workbook(filename=path)
    #所有表格的名稱
    #print(file.get_sheet_names())
    sheets = file.sheetnames
    #拿出一個表格
    sheet = file[sheets[0]]
    #最大行數
    # print(sheet.max_row)
    #最大列數
    # print(sheet.max_column)
    #表名
    # print(sheet.title)

    for lineNum in range(1, sheet.max_row + 1):
        #print(lineNum)
        linList = []
        for columnNum in range(1, sheet.max_column + 1):
            #拿數據
            value = sheet.cell(row=lineNum, column=columnNum).value
            #if value != None:
            linList.append(value)

        print(linList)
    # 讀取一張表的數據
#不能處理Xls文件
path = r"E:\code\python\1.xlsx"
readXlsxFile(path)